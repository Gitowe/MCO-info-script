
# Main file to run the MCO Information Script

import requests
import datetime
import time
import pytz
import io
import os
from tkinter import *
import tkinter as tk
import tkinter.font as tkFont
import ctypes
import pyperclip
import openpyxl
from bs4 import BeautifulSoup
from PIL import Image, ImageTk

from api_functions import *
from global_variables import *

print("\nScript is starting, this might take a couple of seconds...")
print("\nPlayer lists have been fetched and stored.")

# Converts time (in seconds) into other time measurements
def convert_seconds(seconds):
    years = seconds // (365 * 24 * 3600)
    seconds %= 365 * 24 * 3600
    days = seconds // (24 * 3600)
    seconds %= 24 * 3600
    hours = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    seconds %= 60

    return years, days, hours, minutes, seconds

# Converts the unix timestap into a legible date and time
def convert_unix_timestamp(timestamp):
    cst = pytz.timezone('US/Central')
    converted_time = datetime.datetime.fromtimestamp(timestamp, tz=cst)
    return converted_time.strftime('%B %d, %Y - %I:%M:%S %p')

# Converts the timestamp to MM/DD/YYYY format (This is for the excel sheet)
def convert_to_mm_dd_yyyy(timestamp):
    converted_time = datetime.datetime.fromtimestamp(timestamp)
    return converted_time.strftime('%m/%d/%Y')



####################################
# Functions for tkinter operations #
####################################

# Clears the widgets from the window
def clear_ui():
    for widget in window.winfo_children():
        widget.destroy()
        
def clear_player_info():
    for widget in window.winfo_children():
        if isinstance(widget, tk.Label) and widget["text"] != "Enter Minecraft username:":
            widget.destroy()
        elif isinstance(widget, tk.Button) and widget["text"] == "Copy Info":
            widget.destroy()
            
# Initiates getting the player info if the Enter key is pressed
def on_enter(event=None):
    player_info_operation()
    
# Copies the player info into the clipboard
def copy_player_info():
    username = entry.get()
    
    try:
        copy_text = f"{username}\n"
        pyperclip.copy(copy_text)
    except IndexError as e:
        print(f"Error: {e}")
        
def create_default_label(parent, text, header):
    if header:
        header_label = tk.Label(parent, text=text, font=header, bg='#383838', fg='#E1E1E1')  # Adjust color as needed
        return header_label
    else:
        normal_label = tk.Label(parent, text=text, bg='#383838', fg='#E1E1E1')  # Adjust color as needed
        return normal_label
    


#######################################
# Updated Player Lists page functions #
#######################################
        
def update_lists_screen():
    clear_ui()
    
    back_button = tk.Button(window, text="Back", command=show_menu, bg='#545454', fg='#E1E1E1')
    back_button.pack(pady=5, padx=10, anchor='nw')
    
    button = tk.Button(window, text="Refresh Player Lists", command=refresh_player_lists, bg='#545454', fg='#E1E1E1')
    button.pack(pady=10)
    
def refresh_player_lists():
    print("\nRefreshing player lists...")
    
    fetch_and_store_players("https://minecraftonline.com/wiki/Category:Former_staff", "Former Staff", "wiki")
    fetch_and_store_players("https://minecraftonline.com/wiki/Category:God_donor", "God Donor", "wiki")
    fetch_and_store_players("https://minecraftonline.com/cgi-bin/getadminlist.sh", "Admins", "api")
    fetch_and_store_players("https://minecraftonline.com/cgi-bin/getmodlist.sh", "Moderators", "api")
    
    formerstafflist = read_players("Former Staff")
    godlist = read_players("God Donor")
    adminlist = read_players("Admins")
    modlist = read_players("Moderators")
    
    player_list_title = tk.Label(window, text="Player list has been refreshed!", font=tkFont.Font(weight="bold", size=11), bg='#383838', fg='#E1E1E1')
    player_list_title.pack()
    
    print("\nPlayer lists have been refreshed.")
    
    return 0 



###############################
# Update Sheet page functions #
###############################

# Displays the Update Sheet page #
def update_excel_screen():
    global entry_filename
    global entry_sheetname
    
    clear_ui()
    
    back_button = tk.Button(window, text="Back", command=show_menu, bg='#545454', fg='#E1E1E1')
    back_button.pack(pady=5, padx=10, anchor='nw')  # 'nw' stands for northwest, i.e., top-left corner

    label_filename = create_default_label(window, "Enter Excel file name (with extension):", None)
    label_filename.pack(pady=15)

    entry_filename = tk.Entry(window, bg='#545454', fg='#E1E1E1')
    entry_filename.pack(pady=10)

    label_sheetname = create_default_label(window, "Enter sheet name:", None)
    label_sheetname.pack(pady=15)

    entry_sheetname = tk.Entry(window, bg='#545454', fg='#E1E1E1')
    entry_sheetname.pack(pady=10)

    button_update_excel_operation = tk.Button(window, text="Update Excel", command=update_excel_operation, bg='#545454', fg='#E1E1E1')
    button_update_excel_operation.pack()

# Updates the excel sheet with correct information based on the names listed in the excel sheet
def update_excel_operation():
    global entry_filename
    global entry_sheetname
    try:
        # Get the Excel file name and sheet name from the user
        excel_file = entry_filename.get()
        sheet_name = entry_sheetname.get()

        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]

        # Find the column index of "Name"
        name_column_index = None
        for cell in sheet[2]:
            if cell.value == "Name":
                name_column_index = cell.column
                break
            
        name_columns = [col_idx for col_idx, cell in enumerate(sheet[2], start=1) if cell.value == "Name"]

        if not name_columns:
            print("Column 'Name' not found in the specified sheet.")
            return

        # Get all usernames from the "Name" column
        for name_column_index in name_columns:
            
            usernames = [cell[0].value for cell in sheet.iter_rows(min_row=3, min_col=name_column_index, max_col=name_column_index)]
            usernames = list(filter(None, usernames))
            
            print("List of usernames fetched from the column:")
            print(usernames)

            # Find the column index to input the first seen date (2 columns to the right of "Name")
            first_seen_column_index = name_column_index + 2
            banned_column_index = first_seen_column_index + 1

            # Iterate through usernames and update the first seen date column
            for row_index, username in enumerate(usernames, start=3):
                time.sleep(4)
                player_info = exponential_backoff_retry(get_player_info_from_api, username)
                if player_info[0] != "NOTFOUND":
                    join_date = int(player_info[0])
                    formatted_date = convert_unix_timestamp(join_date).split(' - ')[0]  # Extract date only
                    formatted_date = datetime.datetime.strptime(formatted_date, '%B %d, %Y').strftime('%b %d, %Y')
                    
                    # Write the formatted date to the Excel sheet
                    sheet.cell(row=row_index, column=first_seen_column_index).value = formatted_date
                    
                    banned_info = player_info[3].split(";") if len(player_info) > 3 else ["NOTBANNED"]
                    if banned_info[0] != "NOTBANNED":
                        banned_status = "BANNED"
                        sheet.cell(row=row_index, column=banned_column_index).value = banned_status
                        # Print the username being updated along with the ban status
                        print(f"Updated username: {username}, Banned: {banned_status}")

        # Save the changes to the Excel file
        workbook.save(excel_file)
        print("Update completed successfully.")
    except Exception as e:
        print(f"Error occurred: {e}")
        
    
        
##############################
# Server Info page functions #
##############################

def server_info_screen():
    global entry
    
    header_font = tkFont.Font(weight="bold", size=11)
    
    # Clear the previous UI elements
    clear_ui()
    
    back_button = tk.Button(window, text="Back", command=show_menu, bg='#545454', fg='#E1E1E1')
    back_button.pack(pady=5, padx=10, anchor='nw')  

    button = tk.Button(window, text="Players Online", command=player_online_screen, bg='#545454', fg='#E1E1E1')
    button.pack(pady=10)
    
    button = tk.Button(window, text="Server Timeline", command=show_timeline, bg='#545454', fg='#E1E1E1')
    button.pack(pady=10)
    
    print("\nLoading server information, this might take a couple of seconds...")
    
    player_server_info_label = create_default_label(window, f"\nGeneral player information\n", header_font)
    player_server_info_label.pack()
    
    ban_count = get_ban_count_from_api()
    ban_count_label = create_default_label(window, f"Current number of bans is: {ban_count}", None)
    ban_count_label.pack()
    
    unique_visitors = get_unique_visitors_from_api()
    unique_visitors_label = create_default_label(window, f"Current number of unique players is: {unique_visitors}", None)
    unique_visitors_label.pack()
    
    yesterday_visitors = get_yesterday_visitors_from_api()
    yesterday_visitors_label = create_default_label(window, f"Number of unique players yesterday was: {yesterday_visitors}", None)
    yesterday_visitors_label.pack()
    
    unique_visitors_banned = int(ban_count) / int(unique_visitors)
    unique_visitors_banned_label = create_default_label(window, f"Percentage of unique visitors banned: {unique_visitors_banned * 100:.4f}%", None)
    unique_visitors_banned_label.pack()
    
    staff_server_info_label = create_default_label(window, f"\nGeneral staff information\n", header_font)
    staff_server_info_label.pack()
    
    number_of_admins = len(adminlist)
    number_of_admins_label = create_default_label(window, f"Number of admins: {number_of_admins}", None)
    number_of_admins_label.pack()
    
    number_of_mods = len(modlist)
    number_of_mods_label = create_default_label(window, f"Number of mods: {number_of_mods}", None)
    number_of_mods_label.pack()
    
    number_of_formerstaff = len(formerstafflist)
    number_of_formerstaff_label = create_default_label(window, f"Number of former staff: {number_of_formerstaff}", None)
    number_of_formerstaff_label.pack()
    
    number_of_total_staff = number_of_admins + number_of_mods + number_of_formerstaff
    number_of_total_staff_label = create_default_label(window, f"Total number of staff of all time: {number_of_total_staff}", None)
    number_of_total_staff_label.pack()
    
    percent_staff = number_of_total_staff / int(unique_visitors)
    percent_staff_label = create_default_label(window, f"Percentage of player who became staff: {percent_staff * 100:.4f}%", None)
    percent_staff_label.pack()
    
    print("\nServer information has been loaded.")

def player_online_screen():
    
    global mod_color
    global admin_color
    global banned_color
    
    clear_ui()
    
    back_button = tk.Button(window, text="Back", command=server_info_screen, bg='#545454', fg='#E1E1E1')
    back_button.pack(pady=5, padx=10, anchor='nw')  # 'nw' stands for northwest, i.e., top-left corner
   
    player_list = []
    
    get_player_list_from_api(player_list)
    print("\nProcessing player list...")
    
    player_list_title = tk.Label(window, text="Players currently online:", font=tkFont.Font(weight="bold", size=11), bg='#383838', fg='#E1E1E1')
    player_list_title.pack()
   
    if player_list:
        frame = tk.Frame(window, bg='#383838')
        frame.pack(pady=5, padx=10)  # Place the frame in the center of the window
        
        global adminlist  
        global modlist
        
        online_admin_list = sorted([username for username in player_list if adminlist is not None and username in adminlist])
        online_mod_list = sorted([username for username in player_list if modlist is not None and username in modlist])
        online_player_list = sorted([username for username in player_list if username not in modlist and username not in adminlist])
        
        combined_player_list = online_admin_list + online_mod_list + online_player_list
        
        num_admins_label = tk.Label(window, text=f"Admins online: {len(online_admin_list)}", font=tkFont.Font(size=9), fg=admin_color, bg='#383838')
        num_admins_label.pack()
        
        num_mods_label = tk.Label(window, text=f"Mods online: {len(online_mod_list)}", font=tkFont.Font(size=9), fg=mod_color, bg='#383838')
        num_mods_label.pack()
        
        num_players_label = tk.Label(window, text=f"Players online: {len(online_player_list)}", font=tkFont.Font(size=9), fg='#E1E1E1', bg='#383838')
        num_players_label.pack()
    
        print_player_list(combined_player_list, frame)
        
    print("\nPlayer list has been processed.")
        
def get_timeline_from_mediawiki():
    url = "https://minecraftonline.com/wiki/Timeline"
    response = requests.get(url)
    
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, "html.parser")
        timeline = soup.find("div", class_="mw-parser-output")
        if timeline:
            return timeline.get_text()
    return "Failed to fetch timeline data"

def show_timeline():
    timeline_text = get_timeline_from_mediawiki()
    
    if timeline_text:
        popup_window = tk.Toplevel(window)
        popup_window.title("Server Timeline")
        popup_window.geometry("800x400")
        
        canvas_frame = tk.Frame(popup_window)
        canvas_frame.pack(fill=tk.BOTH, expand=True)
        
        canvas = tk.Canvas(canvas_frame, bg="white", width=780, height=380)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar_x = tk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL, command=canvas.xview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        
        canvas.config(xscrollcommand=scrollbar_x.set)
        
        events = timeline_text.split("\n\n")
        num_events = len(events)
        spacing = 700 / num_events  # Adjust the spacing based on the number of events
        
        y_position = 50
        
        for event in events:
            event_info = event.split(" - ")
            if len(event_info) >= 2:
                date = event_info[0]
                description = event_info[1]
                
                canvas.create_text(50, y_position, anchor="w", text=date, font=tkFont.Font(size=10, weight="bold"))
                canvas.create_text(100, y_position, anchor="w", text=description, font=tkFont.Font(size=10))
                canvas.create_line(150, y_position + 5, 750, y_position + 5, fill="black")
                
                y_position += 20  # Adjust vertical position for the next event
        
        canvas.config(scrollregion=canvas.bbox("all"))
        
    else:
        error_label = tk.Label(window, text="Failed to fetch timeline data", font=tkFont.Font(size=9), bg='#383838', fg='red')
        error_label.pack(pady=10, padx=10)




##############################
# Player Info page functions #
##############################

# Displays the Player Info page
def player_info_screen():
    global entry
    
    # Clear the previous UI elements
    clear_ui()
    
    back_button = tk.Button(window, text="Back", command=show_menu, bg='#545454', fg='#E1E1E1')
    back_button.pack(pady=5, padx=10, anchor='nw')  # 'nw' stands for northwest, i.e., top-left corner

    # Create the elements for Player Info screen
    label = create_default_label(window, "Enter Minecraft username:", None)
    label.pack(pady=15)

    entry = tk.Entry(window, bg='#545454', fg='#E1E1E1')
    entry.pack(pady=10)
    
    entry.bind("<Return>", on_enter)
    
    spacer = tk.Label(window, text="")

    button = tk.Button(window, text="Get Info", command=player_info_operation, bg='#545454', fg='#E1E1E1')
    button.pack(pady=10)

# Retrieves player information based on the api
def player_info_operation():
    global entry
    global copy_button
    
    global mod_color
    global admin_color
    global banned_color
    
    header_font = tkFont.Font(weight="bold", size=11)
    
    entry_input = entry.get()
    username = get_real_player_name(entry_input)
    print(username)
    player_info = get_player_info_from_api(username)
    print(player_info)
    
    clear_player_info()
    
    if username == "INVALID":
        player_info_label = create_default_label(window, f"\nInvalid player name.\n", header_font)
        player_info_label.pack()
        return
    elif username == "NOTFOUND":
        player_info_label = create_default_label(window, f"\nPlayer not found.\n", header_font)
        player_info_label.pack()
        return
    
    if player_info[0] == None:
        player_info_label = create_default_label(window, f"\nError in getting player name.\n", header_font)
        player_info_label.pack()
        
    else:
        join_date = int(player_info[0])
        last_seen = int(player_info[1])
        time_seconds = int(player_info[2])
        banned_info = player_info[3].split(";") if len(player_info) > 3 else ["NOTBANNED"]
        
        player_head = get_player_head_from_api(username)
        player_head_photo = ImageTk.PhotoImage(player_head)
        player_head_label = tk.Label(window, image=player_head_photo, bg='#383838')
        player_head_label.player_head = player_head_photo
        player_head_label.pack(pady=10)

        # Check join date and last seen date
        player_info_label = create_default_label(window, f"\nPlayer Information for {username}:", header_font) # Header
        player_info_label.pack()
        
        if username in adminlist:
            username_label = tk.Label(window, text=f"{username} is an administrator", font=tkFont.Font(weight="bold", size=11), fg=admin_color, bg='#383838')
        elif username in modlist:
            username_label = tk.Label(window, text=f"{username} is a moderator", font=tkFont.Font(weight="bold", size=11), fg=mod_color, bg='#383838')
        elif username in formerstafflist:
            username_label = tk.Label(window, text=f"{username} is former staff", font=tkFont.Font(weight="bold", size=11), fg="dark cyan", bg='#383838')
        else:
            username_label = create_default_label(window, f"{username} is not staff", None)
        username_label.pack()

        join_date_label = create_default_label(window, f"Joined: {convert_unix_timestamp(join_date)}", None)
        join_date_label.pack()

        last_seen_label = create_default_label(window, f"Last Seen: {convert_unix_timestamp(last_seen)}", None)
        last_seen_label.pack()

        # Check time played
        years, days, hours, minutes, seconds = convert_seconds(time_seconds)
        if years > 0:
            time_played_label = create_default_label(window, f"Time Played: Years: {years}, Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}", None)
            time_played_label.pack()
        else:
            time_played_label.pack = create_default_label(window, f"Time Played: Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}", None)
            time_played_label.pack()
            
        # Check if player is banned
        ban_info_label = create_default_label(window, f"\nBan Information for {username}:", header_font) # Header
        ban_info_label.pack()

        if banned_info[0] != 'NOTBANNED':
            ban_status_label = tk.Label(window, text=f"Player is banned", fg=banned_color, bg='#383838', font=tkFont.Font(weight="bold", size=11))
            ban_status_label.pack()
        else:
            ban_status_label = create_default_label(window, f"Player is not banned", None)
            ban_status_label.pack()

        if banned_info[0] != "NOTBANNED":
            ban_reason_label = create_default_label(window, f"Ban Reason: {banned_info[2]}", None)
            ban_reason_label.pack()
            
            ban_date_label = create_default_label(window, f"Ban Info: Banned by {banned_info[0]} on {convert_unix_timestamp(int(banned_info[1]))}", None)
            ban_date_label.pack()
            
        calc_info_label = create_default_label(window, f"\nPlayer Calculations for {username}:", header_font) # Header
        calc_info_label.pack()
            
        # Time since join date calculation
        current_time = datetime.datetime.now().timestamp()
        time_since_joined = int(current_time - join_date)
        years, days, hours, minutes, seconds = convert_seconds(time_since_joined)
        if years > 0:
            
            time_since_join_label = create_default_label(window, f"Time Since Joined: Years: {years}, Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}", None)
            time_since_join_label.pack()
        else:
            time_since_join_label = create_default_label(window, f"Time Since Joined: Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}", None)
            time_since_join_label.pack()
        
        # Time since last seen calculation
        time_since_last_seen = int(current_time - last_seen)
        years, days, hours, minutes, seconds = convert_seconds(time_since_last_seen)
        if years > 0:
            time_since_lastseen_label = create_default_label(window, f"Time Since Last Seen: Years: {years}, Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}", None)
            time_since_lastseen_label.pack()
        else:
            time_since_lastseen_label = create_default_label(window, f"Time Since Last Seen: Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}", None)
            time_since_lastseen_label.pack()

        # Time since ban calculation
        if banned_info[0] != "NOTBANNED":
            time_since_ban = int(current_time - int(banned_info[1]))
            years, days, hours, minutes, seconds = convert_seconds(time_since_ban)
            if years > 0:
                time_since_ban_label = create_default_label(window, f"Time Since Ban: Years: {years}, Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}", None)
                time_since_ban_label.pack()
            else:
                time_since_ban_label = create_default_label(window, f"Time Since Ban: Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}", None)
                time_since_ban_label.pack()
                
            
        # Time spent playing since joining calculation
        total_percent_playtime = float(time_seconds / time_since_joined)
        percent_time_since_joining_label = create_default_label(window, f"Percentage of time spent playing since joining: {total_percent_playtime * 100:.4f}%", None)
        percent_time_since_joining_label.pack()
        
        percent_playtime_before_lastseen = float(time_seconds / (time_since_joined - time_since_last_seen))
        percent_time_since_lastseen_label = create_default_label(window, f"Percentage of time spent playing before last seen: {percent_playtime_before_lastseen * 100:.4f}%", None)
        percent_time_since_lastseen_label.pack()
        
        if banned_info[0] != "NOTBANNED":
            percent_playtime_before_ban = float(time_seconds / (time_since_joined - time_since_ban))
            percent_time_since_ban_label = create_default_label(window, f"Percentage of time spent playing before ban: {percent_playtime_before_ban * 100:.4f}%", None)
            percent_time_since_ban_label.pack()
            
        wiki_info = get_player_info_from_wiki(username)
        wiki_info_label = create_default_label(window,f"\nWiki Information for {username}:", header_font) # Header
        wiki_info_label.pack()
        
        if wiki_info:
            for key, value in wiki_info.items():
                if key == "Donor level":
                    wiki_stuff_label = tk.Label(window, text=f"{key}: {value}", bg='#383838', fg='#00AA00')
                elif key == "Kit level":
                    wiki_stuff_label = tk.Label(window, text=f"{key}: {value}", bg='#383838', fg=(check_kit_level(value)))
                elif key == "Legacy donor level":
                    wiki_stuff_label = tk.Label(window, text=f"{key}: {value}", bg='#383838', fg=(check_kit_level(value)))
                else:
                    wiki_stuff_label = create_default_label(window, f"{key}: {value}", None)
                wiki_stuff_label.pack()
                
            wiki_stuff_label = create_default_label(window, f"", None)
            wiki_stuff_label.pack()
        else:
            wiki_stuff_label = create_default_label(window, f"User not found on the wiki\n", None)
            wiki_stuff_label.pack()
        
        copy_button = tk.Button(window, text="Copy Info", command=copy_player_info)
        copy_button.pack(pady=10)
        
def check_kit_level(value):
    color = '#E1E1E1'
    
    if value == "* Wood":
        color = '#fefe3f'
    elif value == "** Stone":
        color = '#fefe3f'
    elif value == "*** Iron":
        color = '#fefe3f'
    elif value == "**** Gold":
        color = '#fefe3f'
    elif value == "***** Diamond":
        color = '#fefe3f'
    elif value == "***** Obsidian":
        color = '#be00be'
    elif value == "***** Nether":
        color = '#be0000'
    elif value == "***** Aether":
        color = '#3ffefe'
    elif value == "***** Demigod":
        color = '#3f3f3f'
    elif value == "*GOD*":
        color = '#fefe3f'
        
    return color

        
############################
# Main Menu page functions #
############################
        
        

# Displays the initial menu
def show_menu():
    # Clear the previous UI elements
    clear_ui()

    # Create the menu buttons
    spacer = tk.Label(window, text="", bg='#383838')
    spacer.pack(pady=50)
    
    player_info_button = tk.Button(window, text="Player Info", command=player_info_screen, bg='#545454', fg='#E1E1E1')
    player_info_button.pack(pady=10, side=TOP)
    
    server_info_button = tk.Button(window, text="Server Info", command=server_info_screen, bg='#545454', fg='#E1E1E1')
    server_info_button.pack(pady=10, side=TOP)

    update_excel_button = tk.Button(window, text="Update Excel", command=update_excel_screen, bg='#545454', fg='#E1E1E1')
    update_excel_button.pack(pady=10, side=TOP)
    
    update_lists_button = tk.Button(window, text="Update Player Lists", command=update_lists_screen, bg='#545454', fg='#E1E1E1')
    update_lists_button.pack(pady=30, side=TOP)
    
    update_lists_button = tk.Button(window, text="Timeline", command=update_lists_screen, bg='#545454', fg='#E1E1E1')
    update_lists_button.pack(pady=30, side=TOP)

# Application window setup stuff
window = tk.Tk()
window.title("Fuzbol's MCO Information Script")

window.geometry("600x650") 
window.configure(bg='#383838')

show_menu()

window.mainloop()