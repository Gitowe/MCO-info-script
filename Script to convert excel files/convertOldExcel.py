from openpyxl import load_workbook
from datetime import datetime

# Function to convert date format
def convert_date_format(date_str):
    try:
        date_obj = datetime.strptime(date_str, '%m/%d/%Y')
        formatted_date = date_obj.strftime('%b %d, %Y')
        return formatted_date
    except ValueError:
        return None  # Return None if not a valid date
    
def clean_text_after_date(text):
    words = text.split('\n')
    cleaned_words = []
    for word in words:
        if word.strip() not in ['MOD', 'ADMIN', 'BANNED', 'FORMER MOD']:
            cleaned_words.append(word.strip())
    return '\n'.join(cleaned_words)

# Load the Excel file
file_path = 'TESTV6.xlsx'  # Replace with your file path
workbook = load_workbook(file_path)
sheet = workbook['TP BREAD']  # Assuming the sheet name is TP TRAILS2

def organize_info(col1, col2, col3):
    # Iterate through rows and modify data
    for row in range(3, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=col1).value
        if cell_value:
            if isinstance(cell_value, str):
                cell_value = cell_value.strip()
            else:
                cell_value = str(cell_value).strip()

            lines = cell_value.split('\n')
            
            
            if len(lines) >= 2:
                username = lines[0]
                date_str = lines[-1]

                # Check if there is a secondary name in parenthesis
                if '(' in lines[1] and ')' in lines[1]:
                    name_parts = lines[1].split('(')
                    name = name_parts[0].strip()
                    name_in_parenthesis = name_parts[1][:-1].strip()
                else:
                    name = lines[1].strip()
                    name_in_parenthesis = None

                # Convert date format
                formatted_date = convert_date_format(date_str)

                cleaned_text = clean_text_after_date(lines[-2])

                # Move data to respective columns
                sheet.cell(row=row, column=col2).value = name_in_parenthesis if name_in_parenthesis else None
                
                sheet.cell(row=row, column=col3).value = formatted_date
                

                # Update Name column and Current Name column
                if name_in_parenthesis:
                    sheet.cell(row=row, column=col1).value = username + '\n' + name
                else:
                    updated_name = username
                    sheet.cell(row=row, column=col1).value = updated_name.strip()

iterations = 5
offset = 4

for i in range(iterations):
    col1, col2, col3 = offset + (i * 6), offset + (i * 6) + 1, offset + (i * 6) + 2
    organize_info(col1, col2, col3)


# Save the updated Excel file
workbook.save('TESTV7.xlsx')
