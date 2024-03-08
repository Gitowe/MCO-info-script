import requests
import datetime
import pytz
import tkinter as tk
import tkinter.font as tkFont
import pyperclip
import openpyxl
from bs4 import BeautifulSoup

entry = None
result_text = None
copy_button = None
entry_filename = None
entry_sheetname = None


# Functions for basic operations


# Converts time (in seconds) into other time measurements
def convert_seconds(seconds):
    years = seconds // (365 * 24 * 3600)
    seconds %= 365 * 24 * 3600
    days = seconds // (24 * 3600)
    seconds %= 24 * 3600
    hours = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    seconds %= 60

    return years, days, hours, minutes, seconds

# Converts the unix timestap into a legible date and time
def convert_unix_timestamp(timestamp):
    cst = pytz.timezone('US/Central')
    converted_time = datetime.datetime.fromtimestamp(timestamp, tz=cst)
    return converted_time.strftime('%B %d, %Y - %I:%M:%S %p')

# Converts the timestamp to MM/DD/YYYY format (This is for the excel sheet)
def convert_to_mm_dd_yyyy(timestamp):
    converted_time = datetime.datetime.fromtimestamp(timestamp)
    return converted_time.strftime('%m/%d/%Y')

# Clears the widgets from the window
def clear_ui():
    for widget in window.winfo_children():
        widget.destroy()

# Gets information from the getplayerinfo portion of the MCO site api
def get_player_info_from_api(username):
    url = f"https://minecraftonline.com/cgi-bin/getplayerinfo?{username}"
    response = requests.get(url)
    if response.status_code == 200:
        return response.text.strip().split("\n")
    else:
        return ["NOTFOUND"]
    
def get_player_info_from_wiki(username):
    user_info = {}
    user_page_url = f"https://minecraftonline.com/wiki/User:{username}"
    
    try:
        response = requests.get(user_page_url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Find the user infobox
            infobox = soup.find("table", class_="infobox")
            if infobox:
                rows = infobox.find_all("tr")
                for row in rows:
                    # Extracting key-value pairs from the infobox
                    cells = row.find_all(["th", "td"])
                    if len(cells) == 2:
                        key = cells[0].text.strip()
                        value = cells[1].text.strip()
                        user_info[key] = value
            
            return user_info
        
        else:
            print(f"Failed to retrieve user information. Error code: {response.status_code}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        
    return None
    
# Initiates getting the player info if the Enter key is pressed
def on_enter(event=None):
    player_info_operation()
    
# Copies the player info into the clipboard
def copy_player_info():
    # Get the player information
    username = entry.get()
    player_info = get_player_info_from_api(username)

    try:
         # Format the information to be copied
        copy_text = f"{username}\n"

        # Copy the formatted information to the clipboard
        pyperclip.copy(copy_text)
    except IndexError as e:
        print(f"Error: {e}")
        
        
# Functions of the pages         
        

# Displays the Update Sheet page
def update_excel_screen():
    global entry_filename
    global entry_sheetname
    
    clear_ui()

    label_filename = tk.Label(window, text="Enter Excel file name (with extension):")
    label_filename.pack()

    entry_filename = tk.Entry(window)
    entry_filename.pack()

    label_sheetname = tk.Label(window, text="Enter sheet name:")
    label_sheetname.pack()

    entry_sheetname = tk.Entry(window)
    entry_sheetname.pack()

    button_update_excel_operation = tk.Button(window, text="Update Excel", command=update_excel_operation)
    button_update_excel_operation.pack()

    button_back = tk.Button(window, text="Back", command=show_menu)
    button_back.pack()

# Updates the excel sheet with correct information based on the names listed in the excel sheet
def update_excel_operation():
    global entry_filename
    global entry_sheetname
    try:
        # Get the Excel file name and sheet name from the user
        excel_file = entry_filename.get()
        sheet_name = entry_sheetname.get()

        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]

        # Find the column index of "Name"
        name_column_index = None
        for cell in sheet[2]:
            if cell.value == "Name":
                name_column_index = cell.column
                break
            
        name_columns = [col_idx for col_idx, cell in enumerate(sheet[2], start=1) if cell.value == "Name"]

        if not name_columns:
            print("Column 'Name' not found in the specified sheet.")
            return

        # Get all usernames from the "Name" column
        for name_column_index in name_columns:
            
            usernames = [cell[0].value for cell in sheet.iter_rows(min_row=3, min_col=name_column_index, max_col=name_column_index)]
            usernames = list(filter(None, usernames))
            
            print("List of usernames fetched from the column:")
            print(usernames)

            # Find the column index to input the first seen date (2 columns to the right of "Name")
            first_seen_column_index = name_column_index + 2
            banned_column_index = first_seen_column_index + 1

            # Iterate through usernames and update the first seen date column
            for row_index, username in enumerate(usernames, start=3):
                player_info = get_player_info_from_api(username)
                if player_info[0] != "NOTFOUND":
                    join_date = int(player_info[0])
                    formatted_date = convert_unix_timestamp(join_date).split(' - ')[0]  # Extract date only
                    formatted_date = datetime.datetime.strptime(formatted_date, '%B %d, %Y').strftime('%b %d, %Y')
                    
                    # Write the formatted date to the Excel sheet
                    sheet.cell(row=row_index, column=first_seen_column_index).value = formatted_date
                    
                    banned_info = player_info[3].split(";") if len(player_info) > 3 else ["NOTBANNED"]
                    if banned_info[0] != "NOTBANNED":
                        banned_status = "BANNED"
                        sheet.cell(row=row_index, column=banned_column_index).value = banned_status
                        # Print the username being updated along with the ban status
                        print(f"Updated username: {username}, Banned: {banned_status}")

        # Save the changes to the Excel file
        workbook.save(excel_file)
        print("Update completed successfully.")
    except Exception as e:
        print(f"Error occurred: {e}")

# Displays the Player Info page
def player_info_screen():
    global entry
    global result_text
    global copy_button
    
    # Clear the previous UI elements
    clear_ui()

    # Create the elements for Player Info screen
    label = tk.Label(window, text="Enter Minecraft username:")
    label.pack(pady=15)

    entry = tk.Entry(window)
    entry.pack(pady=10)
    
    entry.bind("<Return>", on_enter)
    
    spacer = tk.Label(window, text="")
    spacer.pack()

    button = tk.Button(window, text="Get Info", command=player_info_operation)
    button.pack(pady=5)
    
    copy_button = tk.Button(window, text="Copy Info", command=copy_player_info)
    # Hides copy button until player information is found
    copy_button.pack_forget
    
    result_text = tk.StringVar()
    result_label = tk.Label(window, textvariable=result_text)
    result_label.pack()
    
    spacer.pack()

    back_button = tk.Button(window, text="Back", command=show_menu)
    back_button.pack(pady=5)

# Retrieves player information based on the api
def player_info_operation():
    global entry
    global result_text
    global copy_button
    
    username = entry.get()
    player_info = get_player_info_from_api(username)
    
    if player_info[0] == "NOTFOUND":
        result_text.set("Player not found.")
    else:
        join_date = int(player_info[0])
        last_seen = int(player_info[1])
        time_seconds = int(player_info[2])
        banned_info = player_info[3].split(";") if len(player_info) > 3 else ["NOTBANNED"]

        # Check join date and last seen date
        result = f"Player Information for {username}:\n"
        result += f"Joined: {convert_unix_timestamp(join_date)}\n"
        result += f"Last Seen: {convert_unix_timestamp(last_seen)}\n"

        # Check time played
        years, days, hours, minutes, seconds = convert_seconds(time_seconds)
        if years > 0:
            result += f"Time Played: Years: {years}, Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}\n"
        else:
            result += f"Time Played: Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}\n"

        # Check if player is banned
        result += f"\nBan Information for {username}:\n"
        result += f"Ban Status: {'BANNED' if banned_info[0] != 'NOTBANNED' else 'NOT BANNED'}\n"
        if banned_info[0] != "NOTBANNED":
            ban_date = int(banned_info[1])
            result += f"Ban Info: Banned by {banned_info[0]} on {convert_unix_timestamp(ban_date)}\n"
            result += f"Ban Reason: {banned_info[2]}\n"
        result += f"\n"
        
        result += f"Player Calculations for {username}:\n"
        current_time = datetime.datetime.now().timestamp()
        
        # Time since join date calculation
        time_since_joined = int(current_time - join_date)
        years, days, hours, minutes, seconds = convert_seconds(time_since_joined)
        if years > 0:
            result += f"Time Since Joined: Years: {years}, Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}\n"
        else:
            result += f"Time Since Joined: Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}\n"
        
        # Time since last seen calculation
        time_since_last_seen = int(current_time - last_seen)
        years, days, hours, minutes, seconds = convert_seconds(time_since_last_seen)
        if years > 0:
            result += f"Time Since Last Seen: Years: {years}, Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}\n"
        else:
            result += f"Time Since Last Seen: Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}\n"

        # Time since ban calculation
        if banned_info[0] != "NOTBANNED":
            time_since_ban = int(current_time - int(banned_info[1]))
            years, days, hours, minutes, seconds = convert_seconds(time_since_ban)
            if years > 0:
                result += f"Time Since Ban: Years: {years}, Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}\n"
            else:
                result += f"Time Since Ban: Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}\n"
                
        # Time spent playing since joining calculation
        total_percent_playtime = float(time_seconds / time_since_joined)
        result += f"Percentage of time spent playing since joining: {total_percent_playtime * 100:.4f}%\n"
        
        percent_playtime_before_lastseen = float(time_seconds / (time_since_joined - time_since_last_seen))
        result += f"Percentage of time spent playing before last seen: {percent_playtime_before_lastseen * 100:.4f}%\n"
        
        if banned_info[0] != "NOTBANNED":
            percent_playtime_before_ban = float(time_seconds / (time_since_joined - time_since_ban))
            result += f"Percentage of time spent playing before ban: {percent_playtime_before_ban * 100:.4f}%\n"
            
        wiki_info = get_player_info_from_wiki(username)
        result += f"\n"
        result += f"Wiki Information for {username}:\n"
        if wiki_info:
            for key, value in wiki_info.items():
                result += f"{key}: {value}\n"
        else:
            result += f"User not found on the wiki\n"

        copy_button.pack()
        result_text.set(result)

# Displays the initial menu
def show_menu():
    # Clear the previous UI elements
    clear_ui()

    # Create the menu buttons
    player_info_button = tk.Button(window, text="Player Info", command=player_info_screen)
    player_info_button.pack(pady=10)

    update_excel__button = tk.Button(window, text="Update Excel", command=update_excel_screen)
    update_excel__button.pack(pady=10)

# Application window setup stuff
window = tk.Tk()
window.title("MCO Player Information")

window.geometry("600x650") 

show_menu()

window.mainloop()