import requests
import datetime
import pytz
import io
import tkinter as tk
import tkinter.font as tkFont
import pyperclip
import openpyxl
from bs4 import BeautifulSoup
from PIL import Image, ImageTk

entry = None


# I will list some useful api commands here for future reference.
# For some reason there is very little documentation on the wiki or github, so I guess I will store the information here for now.
# Lazy ass devs smgdh. 

# For context, {} - indicates username entry, [] - indicates optional entry, () - indicates non username entry
# Additionally, most (if not all) of the api commands for players should work with uuid as well with _uuid?{name} replacing the end of the command

# 👍 https://minecraftonline.com/cgi-bin/getcorrectname?{name} - Returns the closest name or fixes incorrect capitalization for the name 
# 👍 https://minecraftonline.com/cgi-bin/getplayerinfo?{name} - Returns /firstseen, /lastseen, /timeplayed, and /reason
#       https://minecraftonline.com/cgi-bin/gettimeonline{name} - Returns /timeplayed
#       https://minecraftonline.com/cgi-bin/getfirstseen{name} - Returns /firstseen
#       https://minecraftonline.com/cgi-bin/getlastseen{name} - Returns /lastseen
# 👍 https://minecraftonline.com/cgi-bin/getplayerhead.sh?{name}[&(size of picture).jpg] - Returns picture of the player's head

# https://minecraftonline.com/cgi-bin/getadminlist.sh - Returns a list of admins
# https://minecraftonline.com/cgi-bin/getmodlist.sh - Returns a list of mods
# https://minecraftonline.com/cgi-bin/getplayerlist.sh - Returns a list of players currently on the server
# https://minecraftonline.com/cgi-bin/getbancount.sh - Returns the number of banned players
# https://minecraftonline.com/cgi-bin/getuniquevisitors.py - Returns the number of unique players on the server
# https://minecraftonline.com/cgi-bin/getuniqueyesterday.py - Returns the number of unique players on the server yesterday


##################################
# Functions for basic operations #
##################################


# Converts time (in seconds) into other time measurements
def convert_seconds(seconds):
    years = seconds // (365 * 24 * 3600)
    seconds %= 365 * 24 * 3600
    days = seconds // (24 * 3600)
    seconds %= 24 * 3600
    hours = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    seconds %= 60

    return years, days, hours, minutes, seconds

# Converts the unix timestap into a legible date and time
def convert_unix_timestamp(timestamp):
    cst = pytz.timezone('US/Central')
    converted_time = datetime.datetime.fromtimestamp(timestamp, tz=cst)
    return converted_time.strftime('%B %d, %Y - %I:%M:%S %p')

# Converts the timestamp to MM/DD/YYYY format (This is for the excel sheet)
def convert_to_mm_dd_yyyy(timestamp):
    converted_time = datetime.datetime.fromtimestamp(timestamp)
    return converted_time.strftime('%m/%d/%Y')



####################################
# Functions for tkinter operations #
####################################



# Clears the widgets from the window
def clear_ui():
    for widget in window.winfo_children():
        widget.destroy()
        
def clear_player_info():
    for widget in window.winfo_children():
        if isinstance(widget, tk.Label) and widget["text"] != "Enter Minecraft username:":
            widget.destroy()
        elif isinstance(widget, tk.Button) and widget["text"] == "Copy Info":
            widget.destroy()
            
# Initiates getting the player info if the Enter key is pressed
def on_enter(event=None):
    player_info_operation()
    
# Copies the player info into the clipboard
def copy_player_info():
    # Get the player information
    username = entry.get()
    
    try:
         # Format the information to be copied
        copy_text = f"{username}\n"

        # Copy the formatted information to the clipboard
        pyperclip.copy(copy_text)
    except IndexError as e:
        print(f"Error: {e}")
        
def create_header_label(parent, text):
    header_font = tkFont.Font(weight="bold", size=11)  # Adjust size and weight as needed
    header_label = tk.Label(parent, text=text, font=header_font)
    return header_label

def create_color_label(parent, text, color):
    color_label = tk.Label(parent, text=text, fg=color)  # Adjust color as needed
    return color_label

def create_normal_label(parent, text):
    normal_label = tk.Label(parent, text=text)  # Adjust color as needed
    return normal_label



#######################################
# Functions for api calling functions #
#######################################



def remove_brackets(input_string):
    lines = input_string
    cleaned_lines = [line[2:-2] if line.startswith("['") and line.endswith("']") else line for line in lines]
    
    cleaned_string = "\n".join(cleaned_lines)
    return cleaned_string

def get_real_player_name(username):
    url = f"https://minecraftonline.com/cgi-bin/getcorrectname?{username}"
    response = requests.get(url)
    if response.status_code == 200:
        input_string = response.text.strip().split("\n")
        result = remove_brackets(input_string)
        return result
    else:
        return [None]

def get_player_info_from_api(username):
    url = f"https://minecraftonline.com/cgi-bin/getplayerinfo?{username}"
    response = requests.get(url)
    if response.status_code == 200:
        return response.text.strip().split("\n")
    else:
        return ["NOTFOUND"]
    
def get_player_head_from_api(username):
    url = f"https://minecraftonline.com/cgi-bin/getplayerhead.sh?{username}&64.png"
    response = requests.get(url)
    if response.status_code == 200:
        image = Image.open(io.BytesIO(response.content))
        return image

# Gets information from the MCO wiki (specifically from the userpage)
def get_player_info_from_wiki(username):
    user_info = {}
    user_page_url = f"https://minecraftonline.com/wiki/User:{username}"
    
    try:
        response = requests.get(user_page_url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Find the user infobox
            infobox = soup.find("table", class_="infobox")
            if infobox:
                rows = infobox.find_all("tr")
                for row in rows:
                    # Extracting key-value pairs from the infobox
                    cells = row.find_all(["th", "td"])
                    if len(cells) == 2:
                        key = cells[0].text.strip()
                        value = cells[1].text.strip()
                        user_info[key] = value
            
            return user_info
        
        else:
            print(f"Failed to retrieve user information. Error code: {response.status_code}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        
    return None



###############################
# Update Sheet page functions #
###############################


# Displays the Update Sheet page #

def update_excel_screen():
    global entry_filename
    global entry_sheetname
    
    clear_ui()

    label_filename = tk.Label(window, text="Enter Excel file name (with extension):")
    label_filename.pack()

    entry_filename = tk.Entry(window)
    entry_filename.pack()

    label_sheetname = tk.Label(window, text="Enter sheet name:")
    label_sheetname.pack()

    entry_sheetname = tk.Entry(window)
    entry_sheetname.pack()

    button_update_excel_operation = tk.Button(window, text="Update Excel", command=update_excel_operation)
    button_update_excel_operation.pack()

    button_back = tk.Button(window, text="Back", command=show_menu)
    button_back.pack()

# Updates the excel sheet with correct information based on the names listed in the excel sheet
def update_excel_operation():
    global entry_filename
    global entry_sheetname
    try:
        # Get the Excel file name and sheet name from the user
        excel_file = entry_filename.get()
        sheet_name = entry_sheetname.get()

        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]

        # Find the column index of "Name"
        name_column_index = None
        for cell in sheet[2]:
            if cell.value == "Name":
                name_column_index = cell.column
                break
            
        name_columns = [col_idx for col_idx, cell in enumerate(sheet[2], start=1) if cell.value == "Name"]

        if not name_columns:
            print("Column 'Name' not found in the specified sheet.")
            return

        # Get all usernames from the "Name" column
        for name_column_index in name_columns:
            
            usernames = [cell[0].value for cell in sheet.iter_rows(min_row=3, min_col=name_column_index, max_col=name_column_index)]
            usernames = list(filter(None, usernames))
            
            print("List of usernames fetched from the column:")
            print(usernames)

            # Find the column index to input the first seen date (2 columns to the right of "Name")
            first_seen_column_index = name_column_index + 2
            banned_column_index = first_seen_column_index + 1

            # Iterate through usernames and update the first seen date column
            for row_index, username in enumerate(usernames, start=3):
                player_info = get_player_info_from_api(username)
                if player_info[0] != "NOTFOUND":
                    join_date = int(player_info[0])
                    formatted_date = convert_unix_timestamp(join_date).split(' - ')[0]  # Extract date only
                    formatted_date = datetime.datetime.strptime(formatted_date, '%B %d, %Y').strftime('%b %d, %Y')
                    
                    # Write the formatted date to the Excel sheet
                    sheet.cell(row=row_index, column=first_seen_column_index).value = formatted_date
                    
                    banned_info = player_info[3].split(";") if len(player_info) > 3 else ["NOTBANNED"]
                    if banned_info[0] != "NOTBANNED":
                        banned_status = "BANNED"
                        sheet.cell(row=row_index, column=banned_column_index).value = banned_status
                        # Print the username being updated along with the ban status
                        print(f"Updated username: {username}, Banned: {banned_status}")

        # Save the changes to the Excel file
        workbook.save(excel_file)
        print("Update completed successfully.")
    except Exception as e:
        print(f"Error occurred: {e}")
        
        
        
##############################
# Server Info page functions #
##############################


# Displays the Player Info page
def server_info_screen():
    global entry
    
    # Clear the previous UI elements
    clear_ui()
    
    back_button = tk.Button(window, text="Back", command=show_menu)
    back_button.pack(pady=5, padx=10, anchor='nw')  # 'nw' stands for northwest, i.e., top-left corner

    # Create the elements for Player Info screen
    label = tk.Label(window, text="Enter Minecraft username:")
    label.pack(pady=15)

    entry = tk.Entry(window)
    entry.pack(pady=10)
    
    entry.bind("<Return>", on_enter)
    
    spacer = tk.Label(window, text="")

    button = tk.Button(window, text="Get Info", command=player_info_operation)
    button.pack(pady=10)

# https://minecraftonline.com/cgi-bin/getadminlist.sh - Returns a list of admins
# https://minecraftonline.com/cgi-bin/getmodlist.sh - Returns a list of mods
# https://minecraftonline.com/cgi-bin/getplayerlist.sh - Returns a list of players currently on the server
# https://minecraftonline.com/cgi-bin/getbancount.sh - Returns the number of banned players
# https://minecraftonline.com/cgi-bin/getuniquevisitors.py - Returns the number of unique players on the server

##############################
# Player Info page functions #
##############################



# Displays the Player Info page
def player_info_screen():
    global entry
    
    # Clear the previous UI elements
    clear_ui()
    
    back_button = tk.Button(window, text="Back", command=show_menu)
    back_button.pack(pady=5, padx=10, anchor='nw')  # 'nw' stands for northwest, i.e., top-left corner

    # Create the elements for Player Info screen
    label = tk.Label(window, text="Enter Minecraft username:")
    label.pack(pady=15)

    entry = tk.Entry(window)
    entry.pack(pady=10)
    
    entry.bind("<Return>", on_enter)
    
    spacer = tk.Label(window, text="")

    button = tk.Button(window, text="Get Info", command=player_info_operation)
    button.pack(pady=10)

# Retrieves player information based on the api
def player_info_operation():
    global entry
    global copy_button
    labels_to_pack = []
    
    entry_input = entry.get()
    username = get_real_player_name(entry_input)
    player_info = get_player_info_from_api(username)
    
    clear_player_info()
    
    if player_info[0] == "NOTFOUND":
        player_info = f"\nPlayer not found.\n" # Header
        player_info_label = create_header_label(window, player_info)
        labels_to_pack.append(player_info_label)
        
        for label in labels_to_pack:
                label.pack()
        
    else:
        join_date = int(player_info[0])
        last_seen = int(player_info[1])
        time_seconds = int(player_info[2])
        banned_info = player_info[3].split(";") if len(player_info) > 3 else ["NOTBANNED"]
        
        image = get_player_head_from_api(username)
        photo = ImageTk.PhotoImage(image)
        image_label = tk.Label(window, image=photo)
        image_label.image = photo
        image_label.pack(pady=10)

        # Check join date and last seen date
        player_info = f"\nPlayer Information for {username}:" # Header
        player_info_label = create_header_label(window, player_info)
        labels_to_pack.append(player_info_label)

        join_date_label = create_normal_label(window, f"Joined: {convert_unix_timestamp(join_date)}")
        labels_to_pack.append(join_date_label)

        last_seen_label = create_normal_label(window, f"Last Seen: {convert_unix_timestamp(last_seen)}")
        labels_to_pack.append(last_seen_label)

        # Check time played
        years, days, hours, minutes, seconds = convert_seconds(time_seconds)
        if years > 0:
            time_played_label = create_normal_label(window, f"Time Played: Years: {years}, Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}")
            labels_to_pack.append(time_played_label)
        else:
            time_played_label = create_normal_label(window, f"Time Played: Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}")
            labels_to_pack.append(time_played_label)
            
        # Check if player is banned
        ban_info = f"\nBan Information for {username}:" # Header
        ban_info_label = create_header_label(window, ban_info)
        labels_to_pack.append(ban_info_label)

        ban_status = 'BANNED' if banned_info[0] != 'NOTBANNED' else 'NOT BANNED'
        ban_status_label = create_normal_label(window, f"Ban Status: {ban_status}")
        labels_to_pack.append(ban_status_label)

        if banned_info[0] != "NOTBANNED":
            ban_date_label = create_normal_label(window, f"Ban Info: Banned by {banned_info[0]} on {convert_unix_timestamp(int(banned_info[1]))}")
            labels_to_pack.append(ban_date_label)

            ban_reason_label = create_normal_label(window, f"Ban Reason: {banned_info[2]}")
            labels_to_pack.append(ban_reason_label)
            
        calc_info = f"\nPlayer Calculations for {username}:" # Header
        calc_info_label = create_header_label(window, calc_info)
        labels_to_pack.append(calc_info_label)
            
        # Time since join date calculation
        current_time = datetime.datetime.now().timestamp()
        time_since_joined = int(current_time - join_date)
        years, days, hours, minutes, seconds = convert_seconds(time_since_joined)
        if years > 0:
            
            time_since_join_label = create_normal_label(window, f"Time Since Joined: Years: {years}, Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}")
            labels_to_pack.append(time_since_join_label)
        else:
            time_since_join_label = create_normal_label(window, f"Time Since Joined: Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}")
            labels_to_pack.append(time_since_join_label)
        
        # Time since last seen calculation
        time_since_last_seen = int(current_time - last_seen)
        years, days, hours, minutes, seconds = convert_seconds(time_since_last_seen)
        if years > 0:
            time_since_lastseen_label = create_normal_label(window, f"Time Since Last Seen: Years: {years}, Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}")
            labels_to_pack.append(time_since_lastseen_label)
        else:
            time_since_lastseen_label = create_normal_label(window, f"Time Since Last Seen: Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}")
            labels_to_pack.append(time_since_lastseen_label)

        # Time since ban calculation
        if banned_info[0] != "NOTBANNED":
            time_since_ban = int(current_time - int(banned_info[1]))
            years, days, hours, minutes, seconds = convert_seconds(time_since_ban)
            if years > 0:
                time_since_ban_label = create_normal_label(window, f"Time Since Ban: Years: {years}, Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}")
                labels_to_pack.append(time_since_ban_label)
            else:
                time_since_ban_label = create_normal_label(window, f"Time Since Ban: Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}")
                labels_to_pack.append(time_since_ban_label)
                
            
        # Time spent playing since joining calculation
        total_percent_playtime = float(time_seconds / time_since_joined)
        percent_time_since_joining_label = create_normal_label(window, f"Percentage of time spent playing since joining: {total_percent_playtime * 100:.4f}%")
        labels_to_pack.append(percent_time_since_joining_label)
        
        percent_playtime_before_lastseen = float(time_seconds / (time_since_joined - time_since_last_seen))
        percent_time_since_lastseen_label = create_normal_label(window, f"Percentage of time spent playing before last seen: {percent_playtime_before_lastseen * 100:.4f}%")
        labels_to_pack.append(percent_time_since_lastseen_label)
        
        if banned_info[0] != "NOTBANNED":
            percent_playtime_before_ban = float(time_seconds / (time_since_joined - time_since_ban))
            percent_time_since_ban_label = create_normal_label(window, f"Percentage of time spent playing before ban: {percent_playtime_before_ban * 100:.4f}%")
            labels_to_pack.append(percent_time_since_ban_label)
            
        wiki_info = get_player_info_from_wiki(username)
        wiki_info_label = create_header_label(window, f"\nWiki Information for {username}:")# Header
        labels_to_pack.append(wiki_info_label)
        
        if wiki_info:
            for key, value in wiki_info.items():
                wiki_stuff_label = create_normal_label(window, f"{key}: {value}")
                labels_to_pack.append(wiki_stuff_label)
            wiki_stuff_label = create_normal_label(window, f"")
            labels_to_pack.append(wiki_stuff_label)
        else:
            wiki_stuff_label = create_normal_label(window, f"User not found on the wiki\n")
            labels_to_pack.append(wiki_stuff_label)
    
        for label in labels_to_pack:
            label.pack()
        
        copy_button = tk.Button(window, text="Copy Info", command=copy_player_info)
        copy_button.pack(pady=10)
        
        
        
############################
# Main Menu page functions #
############################
        
        

# Displays the initial menu
def show_menu():
    # Clear the previous UI elements
    clear_ui()

    # Create the menu buttons
    player_info_button = tk.Button(window, text="Player Info", command=player_info_screen)
    player_info_button.pack(pady=10)
    
    server_info_button = tk.Button(window, text="Update Excel", command=server_info_screen)
    server_info_button.pack(pady=10)

    update_excel_button = tk.Button(window, text="Update Excel", command=update_excel_screen)
    update_excel_button.pack(pady=10)

# Application window setup stuff
window = tk.Tk()
window.title("MCO Player Information")

window.geometry("600x650") 

show_menu()

window.mainloop()