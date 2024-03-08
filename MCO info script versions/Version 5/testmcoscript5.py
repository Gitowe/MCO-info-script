import requests
import datetime
import pytz
import io
import os
from tkinter import *
import tkinter as tk
import tkinter.font as tkFont
import ctypes
import pyperclip
import openpyxl
from bs4 import BeautifulSoup
from PIL import Image, ImageTk


# Global variables 

entry = None

banned_color = '#ab3838'
mod_color = '#55FFFF'
admin_color = '#FF5555'
donor_color = '#00AA00'



# I will list some useful api commands here for future reference.
# For some reason there is very little documentation on the wiki or github, so I guess I will store the information here for now.
# Lazy ass devs smgdh. 

# For context, {} - indicates username entry, [] - indicates optional entry, () - indicates non username entry
# Additionally, most (if not all) of the api commands for players should work with uuid as well with _uuid?{name} replacing the end of the command

# 👍 https://minecraftonline.com/cgi-bin/getcorrectname?{name} - Returns the closest name or fixes incorrect capitalization for the name 
# 👍 https://minecraftonline.com/cgi-bin/getplayerinfo?{name} - Returns /firstseen, /lastseen, /timeplayed, and /reason
#       https://minecraftonline.com/cgi-bin/gettimeonline{name} - Returns /timeplayed
#       https://minecraftonline.com/cgi-bin/getfirstseen{name} - Returns /firstseen
#       https://minecraftonline.com/cgi-bin/getlastseen{name} - Returns /lastseen
# 👍 https://minecraftonline.com/cgi-bin/getplayerhead.sh?{name}[&(size of picture).jpg] - Returns picture of the player's head

# 👍 https://minecraftonline.com/cgi-bin/getadminlist.sh - Returns a list of admins
# 👍 https://minecraftonline.com/cgi-bin/getmodlist.sh - Returns a list of mods
# https://minecraftonline.com/cgi-bin/getplayerlist.sh - Returns a list of players currently on the server
# https://minecraftonline.com/cgi-bin/getbancount.sh - Returns the number of banned players
# https://minecraftonline.com/cgi-bin/getuniquevisitors.py - Returns the number of unique players on the server
# https://minecraftonline.com/cgi-bin/getuniqueyesterday.py - Returns the number of unique players on the server yesterday



##################################
# Functions for basic operations #
##################################


# Converts time (in seconds) into other time measurements
def convert_seconds(seconds):
    years = seconds // (365 * 24 * 3600)
    seconds %= 365 * 24 * 3600
    days = seconds // (24 * 3600)
    seconds %= 24 * 3600
    hours = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    seconds %= 60

    return years, days, hours, minutes, seconds

# Converts the unix timestap into a legible date and time
def convert_unix_timestamp(timestamp):
    cst = pytz.timezone('US/Central')
    converted_time = datetime.datetime.fromtimestamp(timestamp, tz=cst)
    return converted_time.strftime('%B %d, %Y - %I:%M:%S %p')

# Converts the timestamp to MM/DD/YYYY format (This is for the excel sheet)
def convert_to_mm_dd_yyyy(timestamp):
    converted_time = datetime.datetime.fromtimestamp(timestamp)
    return converted_time.strftime('%m/%d/%Y')



####################################
# Functions for tkinter operations #
####################################



# Clears the widgets from the window
def clear_ui():
    for widget in window.winfo_children():
        widget.destroy()
        
def clear_player_info():
    for widget in window.winfo_children():
        if isinstance(widget, tk.Label) and widget["text"] != "Enter Minecraft username:":
            widget.destroy()
        elif isinstance(widget, tk.Button) and widget["text"] == "Copy Info":
            widget.destroy()
            
# Initiates getting the player info if the Enter key is pressed
def on_enter(event=None):
    player_info_operation()
    
# Copies the player info into the clipboard
def copy_player_info():
    # Get the player information
    username = entry.get()
    
    try:
         # Format the information to be copied
        copy_text = f"{username}\n"

        # Copy the formatted information to the clipboard
        pyperclip.copy(copy_text)
    except IndexError as e:
        print(f"Error: {e}")
        
def create_default_label(parent, text, header):
    if header:
        header_label = tk.Label(parent, text=text, font=header, bg='#383838', fg='#E1E1E1')  # Adjust color as needed
        return header_label
    else:
        normal_label = tk.Label(parent, text=text, bg='#383838', fg='#E1E1E1')  # Adjust color as needed
        return normal_label
    
    

#######################################
# Functions for api calling functions #
#######################################



def remove_brackets(input_string):
    lines = input_string
    cleaned_lines = [line[2:-2] if line.startswith("['") and line.endswith("']") else line for line in lines]
    
    cleaned_string = "\n".join(cleaned_lines)
    return cleaned_string

def get_real_player_name(username):
    url = f"https://minecraftonline.com/cgi-bin/getcorrectname?{username}"
    response = requests.get(url)
    if response.status_code == 200:
        if response.text.strip() == "NOTFOUND":
            return "NOTFOUND"
        elif response.text.strip() == "INVALID":
            return "INVALID"
        input_string = response.text.strip().split("\n")
        result = remove_brackets(input_string)
        return result
    else:
        return None

def get_player_info_from_api(username):
    url = f"https://minecraftonline.com/cgi-bin/getplayerinfo?{username}"
    response = requests.get(url)
    if response.status_code == 200:
        return response.text.strip().split("\n")
    else:
        return None
    
def get_player_head_from_api(username):
    url = f"https://minecraftonline.com/cgi-bin/getplayerhead.sh?{username}&64.png"
    response = requests.get(url)
    if response.status_code == 200:
        image = Image.open(io.BytesIO(response.content))
        return image
    else:
        return None

# Gets information from the MCO wiki (specifically from the userpage)
def get_player_info_from_wiki(username):
    user_info = {}
    user_page_url = f"https://minecraftonline.com/wiki/User:{username}"
    
    try:
        response = requests.get(user_page_url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Find the user infobox
            infobox = soup.find("table", class_="infobox")
            if infobox:
                rows = infobox.find_all("tr")
                for row in rows:
                    # Extracting key-value pairs from the infobox
                    cells = row.find_all(["th", "td"])
                    if len(cells) == 2:
                        key = cells[0].text.strip()
                        value = cells[1].text.strip()
                        user_info[key] = value
            
            return user_info
        
        else:
            print(f"Failed to retrieve user information. Error code: {response.status_code}")
            return None
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        
    return None



##############################
# Staff list functions/setup #
##############################



def get_player_list_from_api(player_list):
    url = f"https://minecraftonline.com/cgi-bin/getplayerlist.sh"
    response = requests.get(url)
    if response.status_code == 200:
        names = response.text.strip().split("\n")
        player_list.extend
        return player_list
    else:
        return None

mod_file = "modlist.txt"
modlist = []
admin_file = "adminlist.txt"
adminlist = []

def get_staff_list_from_api(url, global_list, file_path):
    
    if os.path.exists(file_path):
        # Read from the existing file
        with open(file_path, "r") as file:
            global_list.extend(file.read().strip().split("\n"))
    else:
        response = requests.get(url)
        if response.status_code == 200:
            names = response.text.strip().split("\n")
            global_list.extend(names)
            # Write the data into the file
            with open(file_path, "w") as file:
                file.write("\n".join(names))
        else:
            print("Failed to retrieve names from the website.")

get_staff_list_from_api("https://minecraftonline.com/cgi-bin/getadminlist.sh", adminlist, mod_file)
get_staff_list_from_api("https://minecraftonline.com/cgi-bin/getmodlist.sh", modlist, admin_file)

def is_user_former_staff(username, file_path):
    # Check if former staff list file exists, if not create one
    
    if not os.path.exists(file_path):
        with open(file_path, "w") as file:
            pass  # Create an empty file

    # Check if the user is listed in the former staff list file
    with open(file_path, "r") as file:
        if username in file.read().split("\n"):
            return True  # User is a former staff

    user_page_url = f"https://minecraftonline.com/wiki/User:{username}"

    try:
        response = requests.get(user_page_url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')

            # Search for the categories section
            categories_section = soup.find("div", class_="mw-normal-catlinks")
            if categories_section:
                # Check if "Former staff" category is present
                categories = categories_section.find_all("a")
                for category in categories:
                    if "Former_staff" in category.get("href", ""):
                        # Add the user to the former staff list file
                        with open(file_path, "a") as file:
                            file.write(username + "\n")
                        return True  # User is a former staff

        else:
            print(f"Failed to retrieve user information. Error code: {response.status_code}")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

    return False



###############################
# Update Sheet page functions #
###############################



# Displays the Update Sheet page #
def update_excel_screen():
    global entry_filename
    global entry_sheetname
    
    clear_ui()
    
    back_button = tk.Button(window, text="Back", command=show_menu, bg='#545454', fg='#E1E1E1')
    back_button.pack(pady=5, padx=10, anchor='nw')  # 'nw' stands for northwest, i.e., top-left corner

    label_filename = create_default_label(window, "Enter Excel file name (with extension):", None)
    label_filename.pack(pady=15)

    entry_filename = tk.Entry(window, bg='#545454', fg='#E1E1E1')
    entry_filename.pack(pady=10)

    label_sheetname = create_default_label(window, "Enter sheet name:", None)
    label_sheetname.pack(pady=15)

    entry_sheetname = tk.Entry(window, bg='#545454', fg='#E1E1E1')
    entry_sheetname.pack(pady=10)

    button_update_excel_operation = tk.Button(window, text="Update Excel", command=update_excel_operation, bg='#545454', fg='#E1E1E1')
    button_update_excel_operation.pack()

# Updates the excel sheet with correct information based on the names listed in the excel sheet
def update_excel_operation():
    global entry_filename
    global entry_sheetname
    try:
        # Get the Excel file name and sheet name from the user
        excel_file = entry_filename.get()
        sheet_name = entry_sheetname.get()

        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]

        # Find the column index of "Name"
        name_column_index = None
        for cell in sheet[2]:
            if cell.value == "Name":
                name_column_index = cell.column
                break
            
        name_columns = [col_idx for col_idx, cell in enumerate(sheet[2], start=1) if cell.value == "Name"]

        if not name_columns:
            print("Column 'Name' not found in the specified sheet.")
            return

        # Get all usernames from the "Name" column
        for name_column_index in name_columns:
            
            usernames = [cell[0].value for cell in sheet.iter_rows(min_row=3, min_col=name_column_index, max_col=name_column_index)]
            usernames = list(filter(None, usernames))
            
            print("List of usernames fetched from the column:")
            print(usernames)

            # Find the column index to input the first seen date (2 columns to the right of "Name")
            first_seen_column_index = name_column_index + 2
            banned_column_index = first_seen_column_index + 1

            # Iterate through usernames and update the first seen date column
            for row_index, username in enumerate(usernames, start=3):
                player_info = get_player_info_from_api(username)
                if player_info[0] != "NOTFOUND":
                    join_date = int(player_info[0])
                    formatted_date = convert_unix_timestamp(join_date).split(' - ')[0]  # Extract date only
                    formatted_date = datetime.datetime.strptime(formatted_date, '%B %d, %Y').strftime('%b %d, %Y')
                    
                    # Write the formatted date to the Excel sheet
                    sheet.cell(row=row_index, column=first_seen_column_index).value = formatted_date
                    
                    banned_info = player_info[3].split(";") if len(player_info) > 3 else ["NOTBANNED"]
                    if banned_info[0] != "NOTBANNED":
                        banned_status = "BANNED"
                        sheet.cell(row=row_index, column=banned_column_index).value = banned_status
                        # Print the username being updated along with the ban status
                        print(f"Updated username: {username}, Banned: {banned_status}")

        # Save the changes to the Excel file
        workbook.save(excel_file)
        print("Update completed successfully.")
    except Exception as e:
        print(f"Error occurred: {e}")
        
        
        
##############################
# Server Info page functions #
##############################



# Displays the Player Info page
def server_info_screen():
    global entry
    
    # Clear the previous UI elements
    clear_ui()
    
    back_button = tk.Button(window, text="Back", command=show_menu, bg='#545454', fg='#E1E1E1')
    back_button.pack(pady=5, padx=10, anchor='nw')  # 'nw' stands for northwest, i.e., top-left corner
   
  
  

    button = tk.Button(window, text="Get Info", command=player_info_operation, bg='#545454', fg='#E1E1E1')
    button.pack(pady=10)
    
    
    
# https://minecraftonline.com/cgi-bin/getplayerlist.sh - Returns a list of players currently on the server
# https://minecraftonline.com/cgi-bin/getbancount.sh - Returns the number of banned players
# https://minecraftonline.com/cgi-bin/getuniquevisitors.py - Returns the number of unique players on the server



##############################
# Player Info page functions #
##############################



# Displays the Player Info page
def player_info_screen():
    global entry
    
    # Clear the previous UI elements
    clear_ui()
    
    back_button = tk.Button(window, text="Back", command=show_menu, bg='#545454', fg='#E1E1E1')
    back_button.pack(pady=5, padx=10, anchor='nw')  # 'nw' stands for northwest, i.e., top-left corner

    # Create the elements for Player Info screen
    label = create_default_label(window, "Enter Minecraft username:", None)
    label.pack(pady=15)

    entry = tk.Entry(window, bg='#545454', fg='#E1E1E1')
    entry.pack(pady=10)
    
    entry.bind("<Return>", on_enter)
    
    spacer = tk.Label(window, text="")

    button = tk.Button(window, text="Get Info", command=player_info_operation, bg='#545454', fg='#E1E1E1')
    button.pack(pady=10)

# Retrieves player information based on the api
def player_info_operation():
    global entry
    global copy_button
    
    global mod_color
    global admin_color
    global banned_color
    
    header_font = tkFont.Font(weight="bold", size=11)
    
    entry_input = entry.get()
    username = get_real_player_name(entry_input)
    print(username)
    player_info = get_player_info_from_api(username)
    print(player_info)
    
    clear_player_info()
    
    if username == "INVALID":
        player_info_label = create_default_label(window, f"\nInvalid player name.\n", header_font)
        player_info_label.pack()
        return
    elif username == "NOTFOUND":
        player_info_label = create_default_label(window, f"\nPlayer not found.\n", header_font)
        player_info_label.pack()
        return
    
    if player_info[0] == None:
        player_info_label = create_default_label(window, f"\nError in getting player name.\n", header_font)
        player_info_label.pack()
        
    else:
        join_date = int(player_info[0])
        last_seen = int(player_info[1])
        time_seconds = int(player_info[2])
        banned_info = player_info[3].split(";") if len(player_info) > 3 else ["NOTBANNED"]
        
        image = get_player_head_from_api(username)
        photo = ImageTk.PhotoImage(image)
        image_label = tk.Label(window, image=photo, bg='#383838')
        image_label.image = photo
        image_label.pack(pady=10)

        # Check join date and last seen date
        player_info_label = create_default_label(window, f"\nPlayer Information for {username}:", header_font) # Header
        player_info_label.pack()
        
        if username in adminlist:
            username_label = tk.Label(window, text=f"{username} is an administrator", font=tkFont.Font(weight="bold", size=11), fg=admin_color, bg='#383838')
        elif username in modlist:
            username_label = tk.Label(window, text=f"{username} is a moderator", font=tkFont.Font(weight="bold", size=11), fg=mod_color, bg='#383838')
        elif is_user_former_staff(username, "formerstaff.txt"):
            username_label = tk.Label(window, text=f"{username} is former staff", font=tkFont.Font(weight="bold", size=11), fg="dark cyan", bg='#383838')
        else:
            username_label = create_default_label(window, f"{username} is not staff", None)
        username_label.pack()

        join_date_label = create_default_label(window, f"Joined: {convert_unix_timestamp(join_date)}", None)
        join_date_label.pack()

        last_seen_label = create_default_label(window, f"Last Seen: {convert_unix_timestamp(last_seen)}", None)
        last_seen_label.pack()

        # Check time played
        years, days, hours, minutes, seconds = convert_seconds(time_seconds)
        if years > 0:
            time_played_label = create_default_label(window, f"Time Played: Years: {years}, Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}", None)
            time_played_label.pack()
        else:
            time_played_label = create_default_label(window, f"Time Played: Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}", None)
            time_played_label.pack()
            
        # Check if player is banned
        ban_info_label = create_default_label(window, f"\nBan Information for {username}:", header_font) # Header
        ban_info_label.pack()

        if banned_info[0] != 'NOTBANNED':
            ban_status_label = tk.Label(window, text=f"Player is banned", fg=banned_color, bg='#383838', font=tkFont.Font(weight="bold", size=11))
            ban_status_label.pack()
        else:
            ban_status_label = create_default_label(window, f"Player is not banned", None)
            ban_status_label.pack()

        if banned_info[0] != "NOTBANNED":
            ban_reason_label = create_default_label(window, f"Ban Reason: {banned_info[2]}", None)
            ban_reason_label.pack()
            
            ban_date_label = create_default_label(window, f"Ban Info: Banned by {banned_info[0]} on {convert_unix_timestamp(int(banned_info[1]))}", None)
            ban_date_label.pack()
            
        calc_info_label = create_default_label(window, f"\nPlayer Calculations for {username}:", header_font) # Header
        calc_info_label.pack()
            
        # Time since join date calculation
        current_time = datetime.datetime.now().timestamp()
        time_since_joined = int(current_time - join_date)
        years, days, hours, minutes, seconds = convert_seconds(time_since_joined)
        if years > 0:
            
            time_since_join_label = create_default_label(window, f"Time Since Joined: Years: {years}, Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}", None)
            time_since_join_label.pack()
        else:
            time_since_join_label = create_default_label(window, f"Time Since Joined: Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}", None)
            time_since_join_label.pack()
        
        # Time since last seen calculation
        time_since_last_seen = int(current_time - last_seen)
        years, days, hours, minutes, seconds = convert_seconds(time_since_last_seen)
        if years > 0:
            time_since_lastseen_label = create_default_label(window, f"Time Since Last Seen: Years: {years}, Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}", None)
            time_since_lastseen_label.pack()
        else:
            time_since_lastseen_label = create_default_label(window, f"Time Since Last Seen: Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}", None)
            time_since_lastseen_label.pack()

        # Time since ban calculation
        if banned_info[0] != "NOTBANNED":
            time_since_ban = int(current_time - int(banned_info[1]))
            years, days, hours, minutes, seconds = convert_seconds(time_since_ban)
            if years > 0:
                time_since_ban_label = create_default_label(window, f"Time Since Ban: Years: {years}, Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}", None)
                time_since_ban_label.pack()
            else:
                time_since_ban_label = create_default_label(window, f"Time Since Ban: Days: {days}, Hours: {hours}, Minutes: {minutes}, Seconds: {seconds}", None)
                time_since_ban_label.pack()
                
            
        # Time spent playing since joining calculation
        total_percent_playtime = float(time_seconds / time_since_joined)
        percent_time_since_joining_label = create_default_label(window, f"Percentage of time spent playing since joining: {total_percent_playtime * 100:.4f}%", None)
        percent_time_since_joining_label.pack()
        
        percent_playtime_before_lastseen = float(time_seconds / (time_since_joined - time_since_last_seen))
        percent_time_since_lastseen_label = create_default_label(window, f"Percentage of time spent playing before last seen: {percent_playtime_before_lastseen * 100:.4f}%", None)
        percent_time_since_lastseen_label.pack()
        
        if banned_info[0] != "NOTBANNED":
            percent_playtime_before_ban = float(time_seconds / (time_since_joined - time_since_ban))
            percent_time_since_ban_label = create_default_label(window, f"Percentage of time spent playing before ban: {percent_playtime_before_ban * 100:.4f}%", None)
            percent_time_since_ban_label.pack()
            
        wiki_info = get_player_info_from_wiki(username)
        wiki_info_label = create_default_label(window,f"\nWiki Information for {username}:", header_font) # Header
        wiki_info_label.pack()
        
        if wiki_info:
            for key, value in wiki_info.items():
                wiki_stuff_label = create_default_label(window, f"{key}: {value}", None)
                wiki_stuff_label.pack()
            wiki_stuff_label = create_default_label(window, f"", None)
            wiki_stuff_label.pack()
        else:
            wiki_stuff_label = create_default_label(window, f"User not found on the wiki\n", None)
            wiki_stuff_label.pack()
        
        copy_button = tk.Button(window, text="Copy Info", command=copy_player_info)
        copy_button.pack(pady=10)
        
        
        
############################
# Main Menu page functions #
############################
        
        

# Displays the initial menu
def show_menu():
    # Clear the previous UI elements
    clear_ui()

    # Create the menu buttons
    spacer = tk.Label(window, text="", bg='#383838')
    spacer.pack(pady=50)
    
    player_info_button = tk.Button(window, text="Player Info", command=player_info_screen, bg='#545454', fg='#E1E1E1')
    player_info_button.pack(pady=10, side=TOP)
    
    server_info_button = tk.Button(window, text="Server Info", command=server_info_screen, bg='#545454', fg='#E1E1E1')
    server_info_button.pack(pady=10, side=TOP)

    update_excel_button = tk.Button(window, text="Update Excel", command=update_excel_screen, bg='#545454', fg='#E1E1E1')
    update_excel_button.pack(pady=10, side=TOP)

# Application window setup stuff
window = tk.Tk()
window.title("MCO Player Information")

window.geometry("600x650") 
window.configure(bg='#383838')

show_menu()

window.mainloop()